from matplotlib import pyplot as plt
import openpyxl as xl

# Load excel file
df=xl.load_workbook('ML_mix.xlsx')
sheet = df['MLmix']

# Load cell values
x_con=[]
y_con=[]
z_con=[]

# row2-MeanDecreaseGini, row3-lift, row4-Importance_1000
for i in range(19):
    c_x_con = sheet.cell(i+2,2)
    c_y_con = sheet.cell(i+2,3)
    c_z_con = sheet.cell(i+2,4)
    x_con.append(c_x_con.value)
    y_con.append(c_y_con.value)
    z_con.append(c_z_con.value)


# Create a figure object
fig1 = plt.figure()
 
# set ax to figure
ax1 = fig1.add_subplot(1, 1, 1)

# Set color map
cm = plt.cm.get_cmap('rainbow') 

# Create bubble chart
scat1 = ax1.scatter(x_con, y_con, c=z_con,s=z_con,alpha=0.5,cmap=cm,vmin=10, vmax=200)


# Add color bar
fig1.colorbar(scat1, ax=ax1)

# Add labels
plt.rcParams['font.family'] = 'Helvetica'
ax1.set_xlabel(sheet.cell(1,5).value)
ax1.set_ylabel(sheet.cell(1,6).value)

ax1.set_xlim(0.3, 1.0) #X axis range
ax1.set_ylim(1.2, 1.7) #Y axis range

ax1.set_xlabel("MeanDecreaseGini[Randomforest]")  #X axis
ax1.set_ylabel("AA[Lift]")           #Y axis
ax1.set_title("Randamforest, XGBoost, MBA")# Title

# Display Graph
plt.show()
plt.savefig("filename.pdf", bbox_inches="tight")

